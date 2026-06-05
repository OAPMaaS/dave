"""
scripts/assign_owners.py — stamp document author metadata with a real team owner.

Why: the chase pipeline resolves a document's owner from its author metadata
(see domain/pipeline.py:_resolve_owner, which substring-matches the configured
owner usernames). Documents with fictional authors match no one and fall back to
the default_owner. This script rewrites the `author` core-property of each file
to a display name that embeds the owner's username as a substring, so
audit_and_persist() routes each document's findings correctly.

Configuration is loaded from config/owner_assignments.json (create it by copying
config/owner_assignments.example.json and filling in your team's data).

Usage:
    python -m scripts.assign_owners            # stamp all files
    python -m scripts.assign_owners --check    # report current vs target, no writes

Idempotent: re-running just re-stamps the same author. Office formats only
(.docx/.pptx/.xlsx + .pdf when pypdf is present). Plain .txt/.csv/.json carry no
author metadata and therefore stay on the default_owner fallback — listed under
UNSTAMPABLE below so the team can decide how to handle them.
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
_CONFIG_FILE = REPO / "config" / "owner_assignments.json"


def _load_config() -> dict:
    """Load owner assignments from config/owner_assignments.json."""
    if not _CONFIG_FILE.exists():
        print(
            f"⚠️  {_CONFIG_FILE} not found.\n"
            f"    Copy config/owner_assignments.example.json and fill in your team data."
        )
        return {"author_names": {}, "owners": {}}
    with open(_CONFIG_FILE) as f:
        return json.load(f)


_cfg = _load_config()

# Display names that embed the platform_username as a substring so
# domain/pipeline.py:_resolve_owner maps author → owner.
AUTHOR_NAME: dict[str, str] = _cfg.get("author_names", {})

# File path (relative to repo root) → owner username
OWNERS: dict[str, str] = _cfg.get("owners", {})

# Files deliberately left unowned to demonstrate the "ownerless document" path
DELIBERATELY_UNOWNED: set[str] = set(_cfg.get("deliberately_unowned", []))

# Formats with no author metadata → _resolve_owner can't read them; they stay on
# the default_owner fallback regardless of the mapping above.
UNSTAMPABLE_EXT = {".txt", ".csv", ".json"}


def _set_author(path: Path, author: str) -> str:
    """Write `author` into the file's metadata. Returns a status string."""
    ext = path.suffix.lower()
    try:
        if ext == ".docx":
            from docx import Document
            d = Document(str(path)); d.core_properties.author = author; d.save(str(path))
        elif ext == ".pptx":
            from pptx import Presentation
            p = Presentation(str(path)); p.core_properties.author = author; p.save(str(path))
        elif ext == ".xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(str(path)); wb.properties.creator = author; wb.save(str(path))
        elif ext == ".pdf":
            try:
                from pypdf import PdfReader, PdfWriter
            except ImportError:
                return "skip (pypdf not installed)"
            r = PdfReader(str(path)); w = PdfWriter(); w.append_pages_from_reader(r)
            w.add_metadata({"/Author": author})
            with open(path, "wb") as fh:
                w.write(fh)
        elif ext in UNSTAMPABLE_EXT:
            return "unstampable (no author metadata)"
        else:
            return f"skip (unsupported {ext})"
    except Exception as exc:  # noqa: BLE001 — report, never crash the batch
        return f"ERROR: {exc}"
    return "ok"


def _current_author(path: Path) -> str:
    ext = path.suffix.lower()
    try:
        if ext == ".docx":
            from docx import Document
            return Document(str(path)).core_properties.author or ""
        if ext == ".pptx":
            from pptx import Presentation
            return Presentation(str(path)).core_properties.author or ""
        if ext == ".xlsx":
            from openpyxl import load_workbook
            return load_workbook(str(path)).properties.creator or ""
        if ext == ".pdf":
            try:
                from pypdf import PdfReader
                return (PdfReader(str(path)).metadata or {}).get("/Author", "") or ""
            except Exception:
                return ""
    except Exception:
        return "?"
    return ""


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--check", action="store_true", help="report only, no writes")
    args = ap.parse_args()

    counts: dict[str, int] = {}
    missing = []
    for rel, owner in OWNERS.items():
        path = REPO / rel
        author = AUTHOR_NAME[owner]
        if not path.exists():
            missing.append(rel)
            continue
        if args.check:
            cur = _current_author(path)
            hit = owner if owner in cur.lower() else "—"
            print(f"  [{owner:7}] {rel}\n            now: {cur!r}  → resolves: {hit}")
        else:
            status = _set_author(path, author)
            counts[status] = counts.get(status, 0) + 1
            print(f"  [{owner:7}] {rel}: {status}")

    print()
    if missing:
        print(f"⚠️  {len(missing)} file(s) in the map not found on disk:")
        for m in missing:
            print(f"     - {m}")
    if not args.check:
        print("Summary:", ", ".join(f"{k}={v}" for k, v in sorted(counts.items())))
    return 0


if __name__ == "__main__":
    sys.exit(main())
